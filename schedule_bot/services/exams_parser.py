from __future__ import annotations

import asyncio
import logging
import re
from dataclasses import dataclass
from io import BytesIO
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _process_workbook_sync(data: bytes) -> bytes:
    """Возвращает копию книги без объединённых ячеек (синхронная версия)."""
    workbook = load_workbook(BytesIO(data))
    changed = False

    for sheet in workbook.worksheets:
        merged_ranges = list(sheet.merged_cells.ranges)
        if not merged_ranges:
            continue
        changed = True
        for merged_range in merged_ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            value = sheet.cell(min_row, min_col).value
            sheet.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(
                    merged_range.min_col,
                    merged_range.max_col + 1,
                ):
                    cell = sheet.cell(row, col)
                    if cell.value is None:
                        cell.value = value

    if not changed:
        return data

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


@dataclass(frozen=True)
class ExamEntry:
    """Запись об экзамене или зачете."""
    date: str
    day_of_week: str
    content: str
    group_name: str


@dataclass(frozen=True)
class ExamSchedule:
    """Расписание экзаменов/зачетов для группы."""
    group_name: str
    entries: List[ExamEntry]


def _normalize_date(date_str: str) -> Tuple[str, str]:
    """
    Извлекает дату и день недели из строки.
    Пример: "26.12.2025\n(пт)" -> ("26.12.2025", "пт")
    """
    if not date_str:
        return "", ""
    
    # Убираем лишние пробелы и переносы строк
    date_str = " ".join(date_str.split())
    
    # Ищем дату в формате dd.mm.yyyy
    date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", date_str)
    if not date_match:
        return "", ""
    
    date = date_match.group(1)
    
    # Ищем день недели в скобках
    day_match = re.search(r"\(([^)]+)\)", date_str)
    day_of_week = day_match.group(1).strip() if day_match else ""
    
    return date, day_of_week


def _extract_exams_schedule_sync(
    data: bytes,
    *,
    sheet_name: str,
    group_name: str,
) -> List[ExamEntry]:
    """
    Извлекает расписание экзаменов из файла с расписанием экзаменов.
    Структура: даты в колонке 2, группы начиная с колонки 3.
    """
    # Обрабатываем объединенные ячейки
    processed_data = _process_workbook_sync(data)
    
    workbook = load_workbook(BytesIO(processed_data), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    
    # Ищем строку с группами (обычно строка 4)
    group_row = None
    group_col = None
    actual_group_name = group_name  # По умолчанию используем исходное имя
    
    # Ищем строку где есть названия групп
    for row_idx in range(1, min(10, sheet.max_row + 1)):
        row_values = [cell.value for cell in sheet[row_idx]]
        # Нормализуем имя группы для поиска
        normalized_target = _normalize_group_name(group_name)
        for col_idx, cell_value in enumerate(row_values, 1):
            if cell_value and isinstance(cell_value, str):
                normalized_cell = _normalize_group_name(cell_value)
                if normalized_cell == normalized_target or normalized_target in normalized_cell:
                    group_row = row_idx
                    group_col = col_idx
                    actual_group_name = str(cell_value).strip()
                    break
        if group_row:
            break
    
    if not group_row or not group_col:
        logger.debug(
            "Group not found in exam schedule sheet=%s target=%s",
            sheet_name,
            group_name,
        )
        return []
    
    # Извлекаем записи
    entries: List[ExamEntry] = []
    
    # Проходим по строкам начиная со строки после группы
    for row_idx in range(group_row + 1, sheet.max_row + 1):
        row_cells = list(sheet[row_idx])
        
        # Ищем дату во второй колонке (индекс 1)
        if len(row_cells) < 2:
            continue
        
        date_cell = row_cells[1].value if len(row_cells) > 1 else None
        if not date_cell:
            continue
        
        date, day_of_week = _normalize_date(str(date_cell))
        if not date:
            continue
        
        # Получаем содержимое для группы
        content_cell = row_cells[group_col - 1].value if len(row_cells) >= group_col else None
        if not content_cell:
            continue
        
        content = str(content_cell).strip()
        if not content or content.lower() in ["", "—", "-", "нет"]:
            continue
        
        entries.append(
            ExamEntry(
                date=date,
                day_of_week=day_of_week,
                content=content,
                group_name=actual_group_name,
            )
        )
    
    workbook.close()
    logger.debug(
        "Extracted exam entries count=%d sheet=%s group=%s",
        len(entries),
        sheet_name,
        group_name,
    )
    return entries


def _extract_credits_schedule_sync(
    data: bytes,
    *,
    sheet_name: str,
    group_name: str,
) -> List[ExamEntry]:
    """
    Извлекает расписание зачетов из файла с расписанием зачетов.
    Структура похожа на обычное расписание: День, Время, затем группы.
    """
    # Обрабатываем объединенные ячейки
    processed_data = _process_workbook_sync(data)
    
    df = pd.read_excel(BytesIO(processed_data), sheet_name=sheet_name, header=6)
    
    # Нормализуем имена колонок
    df = df.rename(columns=_cleanup_column_name)
    
    # Убираем колонки Unnamed
    columns = [
        column
        for column in df.columns
        if not str(column).startswith("Unnamed")
    ]
    df = df[columns]
    
    DAY_COLUMN = "День"
    TIME_COLUMN = "Время занятий"
    
    if DAY_COLUMN not in df.columns or TIME_COLUMN not in df.columns:
        logger.error(
            "Required columns missing in credit schedule sheet=%s columns=%s",
            sheet_name,
            list(df.columns),
        )
        return []
    
    # Ищем группу
    normalized_target = _normalize_group_name(group_name)
    target_group_col = None
    actual_group_name = None
    
    for col in df.columns:
        if col in (DAY_COLUMN, TIME_COLUMN):
            continue
        normalized_col = _normalize_group_name(col)
        if normalized_col == normalized_target or normalized_target in normalized_col:
            target_group_col = col
            actual_group_name = col
            break
    
    if not target_group_col:
        logger.debug(
            "Group not found in credit schedule sheet=%s target=%s available=%s",
            sheet_name,
            group_name,
            [c for c in df.columns if c not in (DAY_COLUMN, TIME_COLUMN)],
        )
        return []
    
    # Нормализуем: заполняем пустые ячейки дня и времени
    df[DAY_COLUMN] = df[DAY_COLUMN].ffill()
    df[TIME_COLUMN] = df[TIME_COLUMN].ffill()
    
    # Удаляем строки где нет данных
    df = df.dropna(subset=[target_group_col], how="all")
    
    # Группируем по дню+время и объединяем содержимое (удаляем дубликаты)
    def combine_unique(series):
        """Объединяет уникальные непустые значения."""
        values = []
        for val in series:
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str and val_str.lower() not in ["", "—", "-", "nan", "none"]:
                    if val_str not in values:
                        values.append(val_str)
        return "\n".join(values) if values else None
    
    grouped = df.groupby([DAY_COLUMN, TIME_COLUMN], as_index=False).agg(
        {target_group_col: combine_unique}
    )
    
    entries: List[ExamEntry] = []
    seen_entries = set()  # Для дополнительной защиты от дубликатов
    
    for _, row in grouped.iterrows():
        day = str(row[DAY_COLUMN]).strip() if pd.notna(row[DAY_COLUMN]) else ""
        time = str(row[TIME_COLUMN]).strip() if pd.notna(row[TIME_COLUMN]) else ""
        content = str(row[target_group_col]).strip() if pd.notna(row[target_group_col]) else ""
        
        if not content or content.lower() in ["", "—", "-", "nan", "none"]:
            continue
        
        # Пропускаем строки-заголовки
        if day.lower() == "день" or time.lower() == "время занятий":
            continue
        
        # Извлекаем дату из содержимого (обычно в конце)
        date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", content)
        if date_match:
            date_str = date_match.group(1)
            # Удаляем дату из содержимого
            content = re.sub(r"\s*\d{2}\.\d{2}\.\d{4}\s*", "", content).strip()
        else:
            # Если даты нет, используем день недели
            date_str = day
        
        # Извлекаем день недели
        day_of_week_match = re.search(r"(понедельник|вторник|среда|четверг|пятница|суббота|воскресенье)", day.lower())
        day_of_week = day_of_week_match.group(1) if day_of_week_match else ""
        
        # Создаем уникальный ключ для проверки дубликатов
        entry_key = (date_str, time, content)
        if entry_key in seen_entries:
            continue
        seen_entries.add(entry_key)
        
        # Формируем полное описание
        full_content = f"{time} — {content}" if time else content
        
        entries.append(
            ExamEntry(
                date=date_str,
                day_of_week=day_of_week,
                content=full_content,
                group_name=actual_group_name or group_name,
            )
        )
    
    logger.debug(
        "Extracted credit entries count=%d sheet=%s group=%s",
        len(entries),
        sheet_name,
        group_name,
    )
    return entries


def _normalize_group_name(name: str) -> str:
    """Нормализует имя группы для сравнения."""
    if not name:
        return ""
    # Удаляем все пробелы, дефисы и приводим к верхнему регистру
    normalized = re.sub(r"\s+|-|—|–", "", str(name).upper())
    # Убираем лишние символы вроде переносов строк
    normalized = normalized.replace("\n", "").strip()
    return normalized


def _cleanup_column_name(name) -> str:
    """Очищает имя колонки."""
    if not isinstance(name, str):
        return str(name) if name else ""
    normalized = re.sub(r"\s+", " ", str(name)).strip()
    if normalized.startswith("Unnamed"):
        return normalized
    # Для кодов групп удаляем лишние пробелы рядом с дефисом
    normalized = normalized.replace(" -", "-").replace("- ", "-")
    normalized = normalized.replace("  ", " ")
    return normalized


async def extract_exams_schedule(
    data: bytes,
    *,
    sheet_name: str,
    group_name: str,
) -> List[ExamEntry]:
    """Асинхронная обертка для извлечения расписания экзаменов."""
    return await asyncio.to_thread(
        _extract_exams_schedule_sync,
        data,
        sheet_name=sheet_name,
        group_name=group_name,
    )


async def extract_credits_schedule(
    data: bytes,
    *,
    sheet_name: str,
    group_name: str,
) -> List[ExamEntry]:
    """Асинхронная обертка для извлечения расписания зачетов."""
    return await asyncio.to_thread(
        _extract_credits_schedule_sync,
        data,
        sheet_name=sheet_name,
        group_name=group_name,
    )

