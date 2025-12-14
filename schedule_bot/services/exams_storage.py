from __future__ import annotations

import logging
import re
from collections import OrderedDict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from schedule_bot.services.exams_parser import ExamEntry, extract_credits_schedule, extract_exams_schedule

logger = logging.getLogger(__name__)


class ExamsStorage:
    """
    Оптимизированное хранилище расписаний зачетов и экзаменов.
    Использует LRU-кэш с ограничением памяти и ленивой загрузкой.
    """
    
    def __init__(
        self,
        exams_dir: Path,
        max_cache_entries: int = 20,  # Максимум 20 групп в кэше
        ttl_minutes: int = 120,  # Время жизни кэша 2 часа
    ) -> None:
        self._exams_dir = Path(exams_dir)
        self._exams_dir.mkdir(parents=True, exist_ok=True)
        
        # LRU кэш: ключ -> (timestamp, данные)
        # OrderedDict сохраняет порядок вставки для LRU
        self._credits_cache: OrderedDict[str, Tuple[datetime, List[ExamEntry]]] = OrderedDict()
        self._exams_cache: OrderedDict[str, Tuple[datetime, List[ExamEntry]]] = OrderedDict()
        
        self._max_cache_entries = max_cache_entries
        self._ttl = timedelta(minutes=ttl_minutes)
        
        # Индекс файлов: тип -> путь к файлу
        self._files_index: Dict[str, Path] = {}
        self._index_files()
        
        logger.debug(
            "ExamsStorage initialised dir=%s max_entries=%d ttl=%d min",
            self._exams_dir,
            max_cache_entries,
            ttl_minutes,
        )
    
    def _index_files(self) -> None:
        """Создает индекс файлов без загрузки в память."""
        for pattern in ["*.xlsx", "*.XLSX"]:
            for file_path in self._exams_dir.glob(pattern):
                file_name = file_path.name.lower()
                
                if "экзамен" in file_name or "экзам" in file_name:
                    self._files_index["exams"] = file_path
                elif "зачет" in file_name or "zachet" in file_name:
                    self._files_index["credits"] = file_path
        
        logger.info(
            "Files indexed: exams=%s credits=%s",
            self._files_index.get("exams"),
            self._files_index.get("credits"),
        )
    
    async def load_all(self) -> None:
        """
        Заглушка для совместимости.
        Теперь загрузка происходит по требованию (lazy loading).
        """
        logger.info("ExamsStorage: using lazy loading, no preload needed")
    
    def _evict_if_needed(self, cache: OrderedDict) -> None:
        """Удаляет старые записи если кэш переполнен."""
        while len(cache) >= self._max_cache_entries:
            # Удаляем самую старую запись (FIFO)
            oldest_key = next(iter(cache))
            del cache[oldest_key]
            logger.debug("Evicted old entry from cache: %s", oldest_key)
    
    def _is_cache_valid(self, timestamp: datetime) -> bool:
        """Проверяет, не истек ли TTL кэша."""
        return datetime.now() - timestamp < self._ttl
    
    async def _load_for_group(
        self,
        group_name: str,
        file_type: str,  # "exams" or "credits"
    ) -> List[ExamEntry]:
        """Загружает расписание для конкретной группы по требованию."""
        file_path = self._files_index.get(file_type)
        if not file_path or not file_path.exists():
            logger.debug("File not found for type=%s", file_type)
            return []
        
        file_content = file_path.read_bytes()
        is_exams = file_type == "exams"
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            # Ищем группу в листах
            for sheet_name in sheet_names:
                if not sheet_name.strip() or sheet_name.lower() == "лист1":
                    continue
                
                try:
                    if is_exams:
                        entries = await extract_exams_schedule(
                            file_content,
                            sheet_name=sheet_name,
                            group_name=group_name,
                        )
                    else:
                        entries = await extract_credits_schedule(
                            file_content,
                            sheet_name=sheet_name,
                            group_name=group_name,
                        )
                    
                    if entries:
                        logger.info(
                            "Loaded %s schedule for group=%s sheet=%s entries=%d",
                            file_type,
                            group_name,
                            sheet_name,
                            len(entries),
                        )
                        return entries
                
                except Exception:
                    logger.debug(
                        "Group not found in sheet=%s type=%s",
                        sheet_name,
                        file_type,
                    )
                    continue
        
        except Exception:
            logger.exception("Failed to load %s for group=%s", file_type, group_name)
        
        return []
    
    
    async def get_credits_for_group(self, group_name: str) -> List[ExamEntry]:
        """
        Возвращает расписание зачетов для группы.
        Использует кэш с lazy loading.
        """
        normalized = self._normalize_group(group_name)
        
        # Проверяем кэш
        if normalized in self._credits_cache:
            timestamp, entries = self._credits_cache[normalized]
            if self._is_cache_valid(timestamp):
                # Перемещаем в конец (LRU)
                self._credits_cache.move_to_end(normalized)
                logger.debug("Cache HIT for credits group=%s", group_name)
                return entries
            else:
                # TTL истек, удаляем
                del self._credits_cache[normalized]
        
        # Загружаем по требованию
        logger.debug("Cache MISS for credits group=%s, loading...", group_name)
        entries = await self._load_for_group(group_name, "credits")
        
        if entries:
            # Эвикция если нужно
            self._evict_if_needed(self._credits_cache)
            # Сохраняем в кэш
            self._credits_cache[normalized] = (datetime.now(), entries)
        
        return entries
    
    async def get_exams_for_group(self, group_name: str) -> List[ExamEntry]:
        """
        Возвращает расписание экзаменов для группы.
        Использует кэш с lazy loading.
        """
        normalized = self._normalize_group(group_name)
        
        # Проверяем кэш
        if normalized in self._exams_cache:
            timestamp, entries = self._exams_cache[normalized]
            if self._is_cache_valid(timestamp):
                # Перемещаем в конец (LRU)
                self._exams_cache.move_to_end(normalized)
                logger.debug("Cache HIT for exams group=%s", group_name)
                return entries
            else:
                # TTL истек, удаляем
                del self._exams_cache[normalized]
        
        # Загружаем по требованию
        logger.debug("Cache MISS for exams group=%s, loading...", group_name)
        entries = await self._load_for_group(group_name, "exams")
        
        if entries:
            # Эвикция если нужно
            self._evict_if_needed(self._exams_cache)
            # Сохраняем в кэш
            self._exams_cache[normalized] = (datetime.now(), entries)
        
        return entries
    
    def _normalize_group(self, name: str) -> str:
        """Нормализует имя группы."""
        if not name:
            return ""
        normalized = re.sub(r"\s+|-|—|–", "", str(name).upper())
        normalized = normalized.replace("\n", "").strip()
        return normalized

