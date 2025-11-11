from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook


DAY_COLUMN = "День"
TIME_COLUMN = "Время занятий"


@dataclass(frozen=True)
class Lesson:
    day: str
    time: str
    description: str


@dataclass(frozen=True)
class LessonBlock:
    title: str
    details: list[str]


def list_sheets(data: bytes) -> List[str]:
    workbook = pd.ExcelFile(BytesIO(data))
    return workbook.sheet_names


def list_groups(data: bytes, sheet_name: str) -> List[str]:
    df = _load_sheet(data, sheet_name)
    excluded = {DAY_COLUMN, TIME_COLUMN}
    return [column for column in df.columns if column not in excluded]


def extract_group_schedule(
    data: bytes,
    *,
    sheet_name: str,
    group_name: str,
    day_filter: Optional[str] = None,
    current_week: Optional[int] = None,
) -> List[Lesson]:
    df = _load_sheet(data, sheet_name)
    if group_name not in df.columns:
        available = ", ".join(df.columns)
        raise ValueError(
            f"Группа '{group_name}' не найдена. Доступные группы: {available}"
        )

    normalized = _normalize_schedule(df)
    lessons: List[Lesson] = []
    for _, row in normalized.iterrows():
        day = row[DAY_COLUMN]
        time = row[TIME_COLUMN]
        content = row[group_name]
        if not isinstance(content, str) or not content.strip():
            continue
        if day_filter and day_filter.lower() != day.lower():
            continue

        filtered_content = content
        if current_week is not None:
            all_blocks = _merge_blocks(content)
            matching_blocks = [
                block
                for block in all_blocks
                if _matches_week(
                    "\n".join([block.title, *block.details]), current_week
                )
            ]
            if not matching_blocks:
                continue
            blocks = matching_blocks
        else:
            blocks = _merge_blocks(filtered_content)
        if not blocks:
            continue

        description = "\n\n".join(_format_block(block) for block in blocks)
        lessons.append(Lesson(day=day, time=time, description=description))
    return lessons


def process_workbook(data: bytes) -> bytes:
    """Возвращает копию книги без объединённых ячеек."""
    workbook = load_workbook(BytesIO(data))
    changed = False

    for sheet in workbook.worksheets:
        merged_ranges = list(sheet.merged_cells.ranges)
        if not merged_ranges:
            continue
        changed = True
        for merged_range in merged_ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            value = sheet.cell(min_row, min_col).value
            sheet.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(
                    merged_range.min_col,
                    merged_range.max_col + 1,
                ):
                    cell = sheet.cell(row, col)
                    if cell.value is None:
                        cell.value = value

    if not changed:
        return data

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def _load_sheet(data: bytes, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(data), sheet_name=sheet_name, header=6)
    df = df.rename(columns=_cleanup_column_name)
    columns = [
        column
        for column in df.columns
        if not column.startswith("Unnamed")
    ]
    df = df[columns]
    return df


def _normalize_schedule(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for column in (DAY_COLUMN, TIME_COLUMN):
        if column in df.columns:
            df[column] = df[column].ffill()
        else:
            raise ValueError(f"В листе нет обязательного столбца '{column}'")

    excluded = {DAY_COLUMN, TIME_COLUMN}
    content_columns = [
        column
        for column in df.columns
        if column not in excluded
    ]

    df = df.dropna(subset=content_columns, how="all")
    aggregated = df.groupby([DAY_COLUMN, TIME_COLUMN], as_index=False).agg(
        {column: _combine_cells for column in content_columns}
    )
    return aggregated


def _combine_cells(series: pd.Series) -> Optional[str]:
    values = [
        str(value).strip()
        for value in series
        if pd.notna(value) and str(value).strip()
    ]
    if not values:
        return None
    # Удаляем дубликаты, сохраняя порядок
    seen = set()
    unique_values = []
    for value in values:
        if value not in seen:
            seen.add(value)
            unique_values.append(value)
    return "\n".join(unique_values)


def _cleanup_column_name(name: str) -> str:
    if not isinstance(name, str):
        return name
    normalized = re.sub(r"\s+", " ", name).strip()
    if normalized.startswith("Unnamed"):
        return normalized
    # Для кодов групп удаляем лишние пробелы рядом с дефисом
    normalized = normalized.replace(" -", "-").replace("- ", "-")
    normalized = normalized.replace("  ", " ")
    return normalized


def _filter_by_week(content: str, time: str, current_week: int) -> str:
    """
    Фильтрует занятия по номеру текущей недели.
    Если в одной ячейке несколько занятий (лекция 6-7н + лаба 8-16н),
    оставляет только подходящие по текущей неделе.
    """
    # Разбиваем контент на строки
    lines = content.split("\n")

    # Пытаемся разбить на отдельные занятия
    # Ищем строки с типом занятия: "(лекции)", "(лаб)", "(практика)" и т.п.
    result_blocks: List[List[str]] = []
    current_lesson = []

    title_markers = [
        "(лек",
        "(лаб",
        "(прак",
        "(сем",
        "ЛЕКЦ",
        "лаб)",
        "прак)",
    ]

    def looks_like_title(text: str) -> bool:
        if text.startswith("http"):
            return False
        lowered = text.lower()
        if "ауд" in lowered:
            return False
        if "." in text:
            return False
        if any(char.isdigit() for char in text[:6]):
            return False
        return True

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        is_title = (
            stripped[0].isupper()
            and (
                any(marker in stripped for marker in title_markers)
                or looks_like_title(stripped)
            )
        )

        if is_title:
            # Проверяем предыдущее занятие
            if current_lesson:
                lesson_text = "\n".join(current_lesson)
                if _matches_week(lesson_text, current_week):
                    result_blocks.append(list(current_lesson))
                current_lesson = []
            current_lesson.append(line)
        else:
            current_lesson.append(line)

    # Проверяем последнее занятие
    if current_lesson:
        lesson_text = "\n".join(current_lesson)
        if _matches_week(lesson_text, current_week):
            result_blocks.append(list(current_lesson))

    if not result_blocks:
        return ""

    return "\n\n".join("\n".join(block) for block in result_blocks)


def _contains_week(text: str) -> bool:
    if not text:
        return False
    return bool(
        re.search(
            r"\d+\s*(?:[-–]\s*\d+)?\s*н",
            text.lower(),
        )
    )


def _merge_blocks(content: str) -> List[LessonBlock]:
    lines = [
        line.strip() for line in content.replace("\r\n", "\n").split("\n")
    ]
    blocks: List[LessonBlock] = []
    title: Optional[str] = None
    details: List[str] = []
    has_week: bool = False

    def is_title(text: str) -> bool:
        if not text:
            return False
        if text.startswith("http"):
            return False
        lowered = text.lower()
        if "ауд" in lowered:
            return False
        if any(char.isdigit() for char in text[:6]):
            return False
        markers = ["(лек", "(лаб", "(прак", "(сем", "(пр", "курс"]
        if text[0].isupper() and any(marker in text for marker in markers):
            return True
        if "." in text:
            return False

        # Если строка с заглавной буквы и не начинается с цифры
        if not text[0].isupper():
            return False

        words = text.split()
        # Длинные названия предметов (3+ слов) без точек - точно название
        if len(words) >= 3:
            return True

        # Одно-два слова с заглавной буквы без точек -
        # тоже может быть названием
        # (например "Электротехника" или "Дискретная математика")
        if len(words) <= 2:
            return True

        return False

    def flush() -> None:
        nonlocal title, details, has_week
        if title:
            blocks.append(LessonBlock(title, details))
        title = None
        details = []
        has_week = False

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        if title is None:
            title = stripped
            has_week = _contains_week(stripped)
            continue

        if is_title(stripped):
            flush()
            title = stripped
            has_week = _contains_week(stripped)
        else:
            if _contains_week(stripped) and has_week:
                flush()
                title = stripped
                has_week = _contains_week(stripped)
                continue
            details.append(stripped)
            if _contains_week(stripped):
                has_week = True

    flush()

    if not blocks:
        stripped = content.strip()
        if not stripped:
            return []
        return [LessonBlock(stripped, [])]

    return blocks


def _format_block(block: LessonBlock) -> str:
    lines = [f"• <b>{block.title}</b>"]
    lines.extend(block.details)
    return "\n".join(lines)


def _matches_week(text: str, current_week: int) -> bool:
    """
    Проверяет, подходит ли текст под текущую неделю.
    Ищет паттерны типа "N-M н", "N н" и т.п.
    Также распознаёт диапазоны без "н" (например "11-14" после "1/2 гр").
    """
    # Убираем явные упоминания подгрупп "1/2 гр" и подобные
    text_clean = re.sub(r"\d+/\d+\s*гр\.?", "", text, flags=re.IGNORECASE)

    # Ищем все упоминания недель в формате "N-M н" или "N н"
    week_patterns_with_n = re.findall(
        r"(\d+)\s*-\s*(\d+)\s*н|\b(\d+)\s*н", text_clean, re.IGNORECASE
    )

    # Ищем диапазоны без "н", которые могут быть неделями
    # (например "11-14" в контексте "1/2 гр 11-14")
    # Ищем паттерн: "гр" затем пробелы и числа N-M
    week_patterns_no_n = re.findall(
        r"гр\.?\s+(\d+)\s*-\s*(\d+)", text, re.IGNORECASE
    )

    all_ranges = []

    # Обрабатываем диапазоны с "н"
    for pattern in week_patterns_with_n:
        if pattern[0] and pattern[1]:  # Диапазон "N-M н"
            start = int(pattern[0])
            end = int(pattern[1])
            all_ranges.append((start, end))
        elif pattern[2]:  # Одиночная неделя "N н"
            week = int(pattern[2])
            all_ranges.append((week, week))

    # Обрабатываем диапазоны без "н" (после "гр")
    for pattern in week_patterns_no_n:
        start = int(pattern[0])
        end = int(pattern[1])
        # Проверяем, что это похоже на недели (обычно < 20)
        if start <= 18 and end <= 18:
            all_ranges.append((start, end))

    if not all_ranges:
        # Если не указаны недели, занятие проходит всегда
        return True

    # Проверяем, попадает ли текущая неделя хотя бы в один диапазон
    for start, end in all_ranges:
        if start <= current_week <= end:
            return True
    return False
